import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Colors
DARK_TEAL = "0F5A47"      # Theme primary (Forest/Agri Green-Teal)
LIGHT_TEAL = "E8F5E9"     # Theme accent (light mint fill for cards)
SUCCESS_GREEN = "2E7D32"  # Forest green for success text
PASSED_BG = "E8F5E9"      # Light mint background for passed cells
PASSED_FG = "2E7D32"      # Dark green for passed text
GRAY_ALT = "F9FAFB"       # Alternating row fill
GRAY_BORDER = "E2E8F0"    # Subtle borders
WHITE = "FFFFFF"
DARK_GRAY = "475569"

def autofit_cols(ws):
    for col in ws.columns:
        if len(col) == 0:
            continue
        max_len = 0
        for cell in col:
            # Skip checking first two rows to avoid overly wide merged title headers
            if cell.row > 2 and cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = "100.0%"
                max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def create_report():
    report_path = "artifacts/test-reports/master-report.json"
    if not os.path.exists(report_path):
        # Fallback to full-e2e-report.json
        report_path = "artifacts/test-reports/full-e2e-report.json"
        
    if not os.path.exists(report_path):
        print(f"Error: Report JSON file not found at {report_path}")
        return
    
    with open(report_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    suites_data = data.get("suites", {})
    
    # Common Style definitions
    font_title = Font(name="Segoe UI", size=18, bold=True, color=WHITE)
    font_section = Font(name="Segoe UI", size=14, bold=True, color=DARK_TEAL)
    font_header = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    font_bold = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    font_regular = Font(name="Segoe UI", size=11, color="334155")
    font_card_val = Font(name="Segoe UI", size=18, bold=True, color=DARK_TEAL)
    font_card_lbl = Font(name="Segoe UI", size=9, bold=True, color=DARK_GRAY)
    
    fill_title = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
    fill_header = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
    fill_alt = PatternFill(start_color=GRAY_ALT, end_color=GRAY_ALT, fill_type="solid")
    fill_card_header = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
    fill_card_body = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin = Side(border_style="thin", color=GRAY_BORDER)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_bottom_double = Border(bottom=Side(border_style="double", color="0F5A47"), top=Side(border_style="thin", color=GRAY_BORDER))
    
    # -------------------------------------------------------------
    # 1. CREATE CONSOLIDATED WORKBOOK
    # -------------------------------------------------------------
    wb_master = openpyxl.Workbook()
    wb_master.remove(wb_master.active) # Remove default sheet
    
    ws_dash = wb_master.create_sheet(title="Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Master Dashboard Title Banner
    ws_dash.merge_cells("A1:G2")
    title_cell = ws_dash["A1"]
    title_cell.value = "AgriConnect Platform - Test Execution Summary Dashboard"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for r in range(1, 3):
        for c in range(1, 8):
            ws_dash.cell(row=r, column=c).fill = fill_title
            
    # Master Dashboard Metadata
    ws_dash["A4"] = "Execution Date:"
    ws_dash["A4"].font = font_bold
    ws_dash["B4"] = data.get("timestamp", "")
    ws_dash["B4"].font = font_regular
    
    ws_dash["D4"] = "Overall Status:"
    ws_dash["D4"].font = font_bold
    ws_dash["E4"] = data.get("overallStatus", "SUCCESS")
    ws_dash["E4"].font = Font(name="Segoe UI", size=11, bold=True, color=SUCCESS_GREEN)
    
    ws_dash["F4"] = "Total Duration:"
    ws_dash["F4"].font = font_bold
    ws_dash["G4"] = f"{data.get('totalDurationSeconds', 0):.2f} seconds"
    ws_dash["G4"].font = font_regular
    
    # Master Dashboard KPI Cards
    total_passed = sum(s.get("passed", 0) for s in suites_data.values())
    total_failed = sum(s.get("failed", 0) for s in suites_data.values())
    kpis = [
        ("TOTAL SUITES", len(suites_data), "A", "A"),
        ("TOTAL TESTS", data.get("totalTests", 0), "B", "C"),
        ("PASSED TESTS", total_passed, "D", "E"),
        ("FAILED TESTS", total_failed, "F", "G")
    ]
    
    for label, val, start_col, end_col in kpis:
        ws_dash.merge_cells(f"{start_col}6:{end_col}6")
        ws_dash.merge_cells(f"{start_col}7:{end_col}7")
        
        lbl_cell = ws_dash[f"{start_col}6"]
        lbl_cell.value = label
        lbl_cell.font = font_card_lbl
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        val_cell = ws_dash[f"{start_col}7"]
        val_cell.value = val
        val_cell.font = font_card_val
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        start_idx = ord(start_col) - ord('A') + 1
        end_idx = ord(end_col) - ord('A') + 1
        for row in range(6, 8):
            for col in range(start_idx, end_idx + 1):
                cell = ws_dash.cell(row=row, column=col)
                cell.border = border_all
                if row == 6:
                    cell.fill = fill_card_header
                else:
                    cell.fill = fill_card_body
                
    # Master Suite Breakdown Table
    ws_dash["A9"] = "Suite Breakdown"
    ws_dash["A9"].font = font_section
    
    headers = ["Suite Code", "Suite Name", "Total Tests", "Passed", "Failed", "Success Rate", "Duration (s)"]
    for idx, h in enumerate(headers, 1):
        cell = ws_dash.cell(row=10, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if idx in [1, 3, 4, 5, 6] else "left", vertical="center")
        cell.border = border_all
        
    current_row = 11
    for code, suite in suites_data.items():
        ws_dash.cell(row=current_row, column=1, value=code).font = font_bold
        ws_dash.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
        
        ws_dash.cell(row=current_row, column=2, value=suite.get("suiteName", "")).font = font_regular
        
        tot = ws_dash.cell(row=current_row, column=3, value=suite.get("totalCount", 0))
        tot.font = font_regular
        tot.alignment = Alignment(horizontal="right")
        
        p = ws_dash.cell(row=current_row, column=4, value=suite.get("passed", 0))
        p.font = font_regular
        p.alignment = Alignment(horizontal="right")
        
        f_val = ws_dash.cell(row=current_row, column=5, value=suite.get("failed", 0))
        f_val.font = font_regular
        f_val.alignment = Alignment(horizontal="right")
        if suite.get("failed", 0) > 0:
            f_val.font = Font(name="Segoe UI", size=11, color="9C0006", bold=True)
            f_val.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
        rate = 1.0 if suite.get("totalCount", 0) == 0 else suite.get("passed", 0) / suite.get("totalCount", 0)
        r = ws_dash.cell(row=current_row, column=6, value=rate)
        r.font = font_regular
        r.number_format = '0.0%'
        r.alignment = Alignment(horizontal="right")
        
        dur = ws_dash.cell(row=current_row, column=7, value=suite.get("durationSeconds", 0.0))
        dur.font = font_regular
        dur.number_format = '0.00'
        dur.alignment = Alignment(horizontal="right")
        
        for col in range(1, 8):
            cell = ws_dash.cell(row=current_row, column=col)
            cell.border = border_all
            if current_row % 2 == 1:
                cell.fill = fill_alt
        current_row += 1
        
    # Master Total Row
    ws_dash.cell(row=current_row, column=1, value="Total").font = font_bold
    ws_dash.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=current_row, column=2, value="").font = font_bold
    
    for col_idx, formula in [
        (3, f"=SUM(C11:C{current_row-1})"),
        (4, f"=SUM(D11:D{current_row-1})"),
        (5, f"=SUM(E11:E{current_row-1})"),
        (6, f"=D{current_row}/C{current_row}"),
        (7, f"=SUM(G11:G{current_row-1})")
    ]:
        cell = ws_dash.cell(row=current_row, column=col_idx, value=formula)
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="right")
        if col_idx == 6:
            cell.number_format = '0.0%'
        elif col_idx == 7:
            cell.number_format = '0.00'
            
    for col in range(1, 8):
        cell = ws_dash.cell(row=current_row, column=col)
        cell.border = border_bottom_double
        cell.fill = fill_card_header
        
    # Chart in Master Dashboard
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Test Execution Duration by Suite (seconds)"
    chart.y_axis.title = "Duration (s)"
    chart.x_axis.title = "Suite"
    
    data_ref = Reference(ws_dash, min_col=7, min_row=10, max_row=current_row-1)
    cats_ref = Reference(ws_dash, min_col=1, min_row=11, max_row=current_row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    chart.height = 10
    chart.width = 16
    ws_dash.add_chart(chart, "A19")
    
    # Autofit master dashboard sheet
    autofit_cols(ws_dash)
    
    # -------------------------------------------------------------
    # 2. CREATE INDIVIDUAL SHEET WORKBOOKS & MASTER DETAILS
    # -------------------------------------------------------------
    clean_filenames = {
        "SEL": "Selenium_Website_Tests",
        "APP": "Appium_Android_Tests",
        "API": "Unit_Tests_API",
        "VAL": "Validation_Tests",
        "DEP": "Deployment_Status",
        "LOD": "Load_Testing_Performance"
    }
    
    for code, suite in suites_data.items():
        tab_name = suite.get("suiteName", code).replace("—", "-").replace("", "-")
        if len(tab_name) > 30:
            tab_name = tab_name[:27] + "..."
            
        # Write to master workbook as a separate tab
        ws_master_tab = wb_master.create_sheet(title=tab_name)
        ws_master_tab.views.sheetView[0].showGridLines = True
        populate_suite_details(ws_master_tab, suite, tab_name, font_title, font_bold, font_regular, font_header, fill_title, fill_header, fill_alt, border_all, SUCCESS_GREEN, PASSED_FG, PASSED_BG)
        autofit_cols(ws_master_tab)
        
        # Create a separate workbook file for this specific sheet
        wb_sep = openpyxl.Workbook()
        wb_sep.remove(wb_sep.active)
        
        # Sep Workbook Sheet 1: Summary Dashboard for this suite
        ws_sep_dash = wb_sep.create_sheet(title="Summary")
        ws_sep_dash.views.sheetView[0].showGridLines = True
        
        # Summary Header Panel
        ws_sep_dash.merge_cells("A1:D2")
        sep_title = ws_sep_dash["A1"]
        sep_title.value = f"Suite Summary: {suite.get('suiteName', tab_name)}"
        sep_title.font = font_title
        sep_title.fill = fill_title
        sep_title.alignment = Alignment(horizontal="center", vertical="center")
        for r in range(1, 3):
            for c in range(1, 5):
                ws_sep_dash.cell(row=r, column=c).fill = fill_title
                
        # Mini Metadata cards
        ws_sep_dash["A4"] = "Suite Code:"
        ws_sep_dash["A4"].font = font_bold
        ws_sep_dash["B4"] = code
        ws_sep_dash["B4"].font = font_regular
        
        ws_sep_dash["C4"] = "Duration (s):"
        ws_sep_dash["C4"].font = font_bold
        ws_sep_dash["D4"] = f"{suite.get('durationSeconds', 0.0):.2f}"
        ws_sep_dash["D4"].font = font_regular
        
        # Metric Boxes
        sep_kpis = [
            ("TOTAL TESTS", suite.get("totalCount", 0), "A", "A"),
            ("PASSED TESTS", suite.get("passed", 0), "B", "B"),
            ("FAILED TESTS", suite.get("failed", 0), "C", "C"),
            ("SUCCESS RATE", f"{(suite.get('passed', 0)/suite.get('totalCount', 1))*100:.1f}%", "D", "D")
        ]
        for label, val, scol, ecol in sep_kpis:
            ws_sep_dash.merge_cells(f"{scol}6:{ecol}6")
            ws_sep_dash.merge_cells(f"{scol}7:{ecol}7")
            
            lcell = ws_sep_dash[f"{scol}6"]
            lcell.value = label
            lcell.font = font_card_lbl
            lcell.alignment = Alignment(horizontal="center", vertical="center")
            lcell.fill = fill_card_header
            
            vcell = ws_sep_dash[f"{scol}7"]
            vcell.value = val
            vcell.font = font_card_val
            vcell.alignment = Alignment(horizontal="center", vertical="center")
            vcell.fill = fill_card_body
            
            ws_sep_dash[f"{scol}6"].border = border_all
            ws_sep_dash[f"{scol}7"].border = border_all
            
        autofit_cols(ws_sep_dash)
        
        # Sep Workbook Sheet 2: Detail view
        ws_sep_det = wb_sep.create_sheet(title="Test Cases Details")
        ws_sep_det.views.sheetView[0].showGridLines = True
        populate_suite_details(ws_sep_det, suite, tab_name, font_title, font_bold, font_regular, font_header, fill_title, fill_header, fill_alt, border_all, SUCCESS_GREEN, PASSED_FG, PASSED_BG)
        autofit_cols(ws_sep_det)
        
        # Save separate file
        filename = clean_filenames.get(code, f"Suite_{code}")
        os.makedirs("artifacts/test-reports/separate-sheets", exist_ok=True)
        out_sep = f"artifacts/test-reports/separate-sheets/{filename}.xlsx"
        wb_sep.save(out_sep)
        print(f"Saved: {os.path.abspath(out_sep)}")
        
    # Save master workbook
    os.makedirs("artifacts/test-reports", exist_ok=True)
    out_master = "artifacts/test-reports/AgriConnect_Test_Report.xlsx"
    wb_master.save(out_master)
    print(f"Saved consolidated master spreadsheet to: {os.path.abspath(out_master)}")

def populate_suite_details(ws, suite, tab_name, font_title, font_bold, font_regular, font_header, fill_title, fill_header, fill_alt, border_all, SUCCESS_GREEN, PASSED_FG, PASSED_BG):
    # Title Banner
    ws.merge_cells("A1:D2")
    suite_title = ws["A1"]
    suite_title.value = f"Suite Details: {suite.get('suiteName', tab_name)}"
    suite_title.font = font_title
    suite_title.fill = fill_title
    suite_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    for r in range(1, 3):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill_title
            
    # Stats summary row (Row 4)
    ws["A4"] = "Total Tests:"
    ws["A4"].font = font_bold
    ws["B4"] = suite.get("totalCount", 0)
    ws["B4"].font = font_regular
    
    ws["C4"] = "Status:"
    ws["C4"].font = font_bold
    ws["D4"] = "PASSED" if suite.get("failed", 0) == 0 else f"FAILED ({suite.get('failed')} failed)"
    ws["D4"].font = Font(name="Segoe UI", size=11, bold=True, color=SUCCESS_GREEN if suite.get("failed", 0) == 0 else "B91C1C")
    
    # Table Headers (Row 6)
    ws_headers = ["Test ID", "Test Case Name", "Status", "Duration (ms)"]
    for idx, h in enumerate(ws_headers, 1):
        cell = ws.cell(row=6, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if idx in [1, 3] else ("right" if idx == 4 else "left"), vertical="center")
        cell.border = border_all
        
    # Data Rows
    row_idx = 7
    for test in suite.get("tests", []):
        tid = ws.cell(row=row_idx, column=1, value=test.get("id", ""))
        tid.font = font_bold
        tid.alignment = Alignment(horizontal="center")
        
        tname = ws.cell(row=row_idx, column=2, value=test.get("name", ""))
        tname.font = font_regular
        
        tstat = ws.cell(row=row_idx, column=3, value=test.get("status", ""))
        tstat.font = Font(name="Segoe UI", size=11, bold=True, color=PASSED_FG)
        tstat.alignment = Alignment(horizontal="center")
        tstat.fill = PatternFill(start_color=PASSED_BG, end_color=PASSED_BG, fill_type="solid")
        
        tdur = ws.cell(row=row_idx, column=4, value=test.get("durationMs", 0.0))
        tdur.font = font_regular
        tdur.alignment = Alignment(horizontal="right")
        tdur.number_format = '#,##0.0'
        
        # Apply border & zebra stripes
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border_all
            if col != 3:
                if row_idx % 2 == 1:
                    cell.fill = fill_alt
                    
        row_idx += 1

if __name__ == '__main__':
    create_report()
